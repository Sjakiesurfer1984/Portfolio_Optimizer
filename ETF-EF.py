import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import mplcursors
from pprint import pprint
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.drawing.image import Image
import os

# Ensure that yfinance is installed in your environment
# and be aware that Yfinance  API is a third-party library
# that may have limitations or changes in the future. Hence, if a data error occurs,
# it is recommended to check the Yfinance documentation or GitHub repository for updates.
class ETFDataHandler:
    def __init__(self, symbols: list[str], start_date: str, end_date: str) -> None:
        """
        Initialize ETFDataHandler with ETF symbols and date range.

        Args:
            symbols (list[str]): List of ticker symbols (e.g., ['AAPL', 'MSFT']).
            start_date (str): Start date in 'YYYY-MM-DD' format.
            end_date (str): End date in 'YYYY-MM-DD' format.
        """
        self.symbols: list[str] = symbols
        self.start_date: str = start_date
        self.end_date: str = end_date
        self.data: pd.DataFrame = self.download_data()

    def download_data(self) -> pd.DataFrame:
        """
        Download adjusted close data for each ticker symbol.

        Returns:
            pd.DataFrame: Adjusted closing prices with datetime index.
        """
        print("starting data download...")

        # Download all columns (including 'Adj Close') for the given symbols and date range
        raw_data = yf.download(
            self.symbols,
            start=self.start_date,
            end=self.end_date,
            auto_adjust=False,  # Keep 'Adj Close' explicitly (not merged with 'Close')
            group_by='column'   # Prevents nesting like ['Price']['AAPL']['Close']
        )

        print(f"Raw data: {raw_data.head()}")  # Show structure of retrieved data

        # Ensure valid response and required column exists
        if raw_data.empty or 'Adj Close' not in raw_data:
            raise ValueError("Failed to retrieve data or missing 'Adj Close'.")

        # Extract just the 'Adj Close' prices
        data = raw_data['Adj Close'].dropna(axis=1, how='all')  # Drop ETFs with only NaNs

        # Show what was successfully downloaded
        print("Downloaded columns:", data.columns.tolist())
        if data.empty:
            raise ValueError("All tickers failed to download any data.")

        # Set self.data early so it's available if needed
        self.data = data

        print(self.data.head())  # Preview of cleaned data
        print("Data downloaded successfully.")

        # Check and synchronize earliest available start date across tickers
        self.verify_shortest_time_range(self.data)

        return self.data

    def verify_shortest_time_range(self, data: pd.DataFrame) -> None:
        """
        Ensure that all ETFs in the dataset have data starting from the same date.
        If not, truncates all ETFs to start from the latest available date.

        Args:
            data (pd.DataFrame): The cleaned adjusted close data.
        """
        print("Verifying shortest time range available for each ticker symbol...")

        # Find the first valid data point for each ETF
        min_dates = data.apply(lambda x: x.first_valid_index())

        # Determine the latest of all those first dates (worst overlap)
        min_date = min_dates.max()

        # If we need to shift the start date to align all tickers
        if min_date and pd.Timestamp(min_date) > pd.Timestamp(self.start_date):
            print(
                f"Adjusting start date from {self.start_date} to {min_date} to ensure equal time range for all ETFs.")
            self.start_date = min_date

            # Drop early data from all ETFs so time ranges match
            self.data = data[data.index >= min_date]
        else:
            # If already aligned, nothing to do
            self.data = data


class ETFMetrics:
    def __init__(self, etf_data_handler: object) -> None:
        """
        Initialize the ETFMetrics object using data from ETFDataHandler.

        Parameters:
            etf_data_handler (object): An instance of ETFDataHandler containing historical ETF price data.
        """
        self.etf_data: pd.DataFrame = etf_data_handler.data                            # Raw price data
        self.daily_returns: pd.DataFrame = self.calculate_daily_returns()              # Daily percentage returns
        self.metrics: pd.DataFrame = self.calculate_individual_metrics()               # Key stats per ETF
        self.correlation_stats: dict = self.calculate_correlation_statistics()         # Correlation matrix & summary

    def calculate_daily_returns(self) -> pd.DataFrame:
        """
        Calculate and clean daily percentage returns for each ETF.

        Returns:
            pd.DataFrame: Daily return matrix with tickers as columns and dates as index.
        """
        # Compute % change between days
        daily_returns = self.etf_data.pct_change().dropna()

        # Convert all values to numeric (safety) and drop any rows with NaNs
        daily_returns = daily_returns.apply(pd.to_numeric, errors='coerce').dropna()

        # Optional debug print for inspection
        print(f"Daily returns: \n{daily_returns.head()}")

        return daily_returns

    def calculate_individual_metrics(self) -> pd.DataFrame:
        """
        Compute per-ETF metrics such as max/min daily return, volatility, total return, CAGR, and Sharpe ratio.

        Returns:
            pd.DataFrame: Metrics summary for each ETF.
        """
        metrics = pd.DataFrame(index=self.etf_data.columns)

        # Max daily return over the whole period
        metrics['Max Return'] = self.daily_returns.max()

        # Min daily return over the whole period
        metrics['Min Return'] = self.daily_returns.min()

        # Annualized volatility: std dev of daily returns × sqrt(252 trading days)
        metrics['Standard Deviation'] = self.daily_returns.std() * np.sqrt(252)

        # Total (cumulative) return over the period
        metrics['Real Return'] = (self.etf_data.iloc[-1] / self.etf_data.iloc[0]) - 1

        # CAGR: convert real return into annualized return based on trading period length
        metrics['Annualized Return'] = (1 + metrics['Real Return']) ** (252 / len(self.daily_returns)) - 1

        # Sharpe Ratio: Annualized Return / Annualized Risk (no risk-free rate subtracted here)
        metrics['Sharpe Ratio'] = metrics['Annualized Return'] / metrics['Standard Deviation']

        return metrics

    def calculate_correlation_statistics(self) -> dict:
        """
        Generate correlation metrics between ETFs.

        Returns:
            dict: Includes static correlation matrix, rolling 30-day correlation,
                  mean correlation, and median correlation across all ETF pairs.
        """
        # Compute pairwise Pearson correlation between ETFs
        correlation_matrix = self.daily_returns.corr()

        # Compute rolling 30-day window correlations
        rolling_correlation = self.daily_returns.rolling(window=30).corr().dropna()

        # Summary values for correlation matrix
        correlation_summary = {
            'Correlation Matrix': correlation_matrix,
            'Mean Correlation': correlation_matrix.mean().mean(),
            'Median Correlation': correlation_matrix.stack().median(),
            'Rolling Correlation (30-day)': rolling_correlation
        }

        return correlation_summary

    def display_correlation_statistics(self) -> None:
        """
        Print the correlation matrix and summary statistics to console.
        """
        print("\nCorrelation Matrix:")
        print(self.correlation_stats['Correlation Matrix'])

        print("\nMean Correlation:", self.correlation_stats['Mean Correlation'])
        print("Median Correlation:", self.correlation_stats['Median Correlation'])

        print("\nRolling Correlation (30-day):")
        print(self.correlation_stats['Rolling Correlation (30-day)'].head())




class PortfolioOptimizer:
    def __init__(self, etf_metrics, risk_free_rate=0.03):
        # Ensure max_etf_weight is feasible for the number of ETFs
        self.num_symbols = len(etf_metrics.daily_returns.columns)
        self.max_etf_weight = 0.3  # Maximum weight for a single ETF
        self.weight_step_size = 0.1  # Weight step size in increments of 1%
        min_required_max_weight = 1.0 / self.num_symbols * 1.2
        if self.max_etf_weight < min_required_max_weight:
            print(
                f"Adjusting max_etf_weight from {self.max_etf_weight} to {min_required_max_weight} to ensure feasibility.")
            self.max_etf_weight = min_required_max_weight
        self.daily_returns = etf_metrics.daily_returns
        self.num_symbols = len(self.daily_returns.columns)
        self.risk_free_rate = risk_free_rate
        self.results_matrix, self.optimal_weights, self.optimal_metrics, self.all_weights = self.optimize_portfolio(
            num_portfolios=10000)


    def generate_random_weights(self, num_symbols: int) -> np.ndarray | None:
        '''
            Generate a random weight vector using a Dirichlet distribution,
            ensuring that no individual weight exceeds the max_etf_weight constraint.
            
            This function attempts random generation for up to 10 seconds. If no
            valid weight vector is found within that time, it returns None.

            Parameters:
                num_symbols (int): Number of assets (ETFs) in the portfolio.

            Returns:
                np.ndarray | None: A valid array of weights summing to 1, or None if timeout is reached.
        '''
        start_time = time.time()
        while True:
            if time.time() - start_time > 10:
                print("Exiting weight generation due to timeout.")
                return None

            weights = np.random.dirichlet(np.ones(num_symbols), size=1)[0]

            if np.all(weights <= self.max_etf_weight):
                return weights


    def calculate_portfolio_performance(self, weights: np.ndarray) -> tuple[float, float, float]:
        """
        Calculate portfolio performance metrics: cumulative return, volatility, and Sharpe ratio.

        Parameters:
            weights (np.ndarray): Array of asset weights. Must sum to 1. These represent the 
                                proportional allocation to each ETF in the portfolio.

        Returns:
            tuple[float, float, float]: A tuple containing:
                - portfolio_cumulative_return: The total return of the portfolio over the full time period.
                - portfolio_std_dev: Annualized portfolio volatility (standard deviation of returns).
                - sharpe_ratio: Risk-adjusted return, defined as excess return per unit of risk.
        """

        # Multiply each day's returns by their corresponding weight and sum — dot product gives
        # the daily return of the whole portfolio over time (vectorized: faster & cleaner).
        # self.daily_returns is a 2D matrix with shape (n_days, n_assets)
        # weights is a 1D vector with shape (n_assets,)
        # So the result is a vector of daily portfolio returns — one value per day.
        weighted_daily_returns = np.dot(self.daily_returns, weights)

        # Compute cumulative return by compounding daily returns:
        # (1 + r1)(1 + r2)... - 1 = total return over the period.
        portfolio_cumulative_return = (1 + weighted_daily_returns).cumprod()[-1] - 1

        # Portfolio volatility (risk) = standard deviation of daily returns,
        # annualized by multiplying with sqrt(252) (252 trading days per year).
        portfolio_std_dev = np.std(weighted_daily_returns) * np.sqrt(252)

        # Sharpe Ratio = (Portfolio Return - Risk-Free Rate) / Volatility.
        # It tells you how much *excess* return you're getting for each unit of risk.
        sharpe_ratio = (portfolio_cumulative_return - self.risk_free_rate) / portfolio_std_dev

        # Return all 3 metrics for use in optimization or reporting.
        return weighted_daily_returns, portfolio_cumulative_return, portfolio_std_dev, sharpe_ratio


    def optimize_portfolio(self, num_portfolios: int) -> tuple[np.ndarray, np.ndarray, dict, list[np.ndarray]]:
        """
        Simulate many random portfolios and select the one with the highest Sharpe ratio.

        This function generates `num_portfolios` random portfolios, each with a random 
        set of weights (subject to constraints), calculates performance metrics for each, 
        and identifies the portfolio with the highest Sharpe ratio.

        Parameters:
            num_portfolios (int): Number of portfolio simulations to run.

        Returns:
            tuple:
                - results_matrix (np.ndarray): 2D array storing returns, risks, Sharpe ratios, and weights.
                - optimal_weights (np.ndarray): The weight vector of the best (Sharpe-optimal) portfolio.
                - optimal_metrics (dict): Dictionary with metrics for the best portfolio.
                - all_weights (list[np.ndarray]): List of all portfolio weight vectors.
        """

        print(f"\nSimulating {num_portfolios} portfolios. This may take a while.")

        # Create a results matrix:
        # Row 0 = cumulative return
        # Row 1 = standard deviation (volatility)
        # Row 2 = Sharpe ratio
        # Rows 3+ = weights for each ETF
        results_matrix = np.zeros((3 + self.num_symbols, num_portfolios))

        all_weights = []            # Stores all weight vectors
        optimal_weights = None      # Tracks weights of the best Sharpe portfolio
        max_sharpe_ratio = -np.inf  # Init to negative infinity for comparison
        optimal_metrics = {}        # Stores metrics of the best portfolio

        for i in range(num_portfolios):
            # Generate a random valid weight vector
            weights = self.generate_random_weights(self.num_symbols)

            # Calculate performance metrics for the generated weights
            weighted_daily_returns, portfolio_cumulative_return, portfolio_std_dev, sharpe_ratio = self.calculate_portfolio_performance(weights)

            # Store the metrics in the matrix
            results_matrix[0, i] = portfolio_cumulative_return
            results_matrix[1, i] = portfolio_std_dev
            results_matrix[2, i] = sharpe_ratio

            # Store the weights themselves
            all_weights.append(weights)
            results_matrix[3:3 + self.num_symbols, i] = weights

            # If this portfolio has a better Sharpe ratio, record it as the best so far
            if sharpe_ratio > max_sharpe_ratio:
                max_sharpe_ratio = sharpe_ratio
                optimal_weights = weights

                # Construct a dictionary of key metrics for the best portfolio
                optimal_metrics = {
                    'Max Return': np.max(weighted_daily_returns),  # highest total daily return in any single day
                    'Min Return': np.min(weighted_daily_returns),  # lowest total daily return in any single day
                    'Standard Deviation': portfolio_std_dev,
                    'Real Return': portfolio_cumulative_return,
                    'Annualized Return': (1 + portfolio_cumulative_return) ** (1 / (len(self.daily_returns) / 252)) - 1,
                    'Sharpe Ratio': sharpe_ratio,
                    'Optimal Weights': list(optimal_weights)
                }

        # Return all results for plotting/reporting
        return results_matrix, optimal_weights, optimal_metrics, all_weights


class EfficientFrontierPlotter:
    def __init__(self, optimizer: object):
        """
        Initialize the plotter with results from the PortfolioOptimizer.

        Parameters:
            optimizer (object): Instance of PortfolioOptimizer containing
                                simulation results and optimal portfolio metrics.
        """
        self.results = optimizer.results_matrix         # Matrix with performance metrics and weights
        self.weights = optimizer.all_weights            # All simulated portfolio weight vectors
        self.optimal_weights = optimizer.optimal_weights
        self.optimal_metrics = optimizer.optimal_metrics
        self.daily_returns = optimizer.daily_returns

    def plot_efficient_frontier(self, save_path: str = 'efficient_frontier.png') -> None:
        """
        Plot the efficient frontier, showing risk vs return for all simulated portfolios,
        highlighting the optimal Sharpe ratio and least-risk portfolios.

        Parameters:
            save_path (str): File path to save the generated PNG plot.
        """
        # Get return (y) and risk (x) of the Sharpe-optimal portfolio
        optimal_return = self.optimal_metrics['Max Return'] + 1  # Convert from growth rate to multiplier (e.g. 0.45 → 1.45)
        optimal_risk = self.optimal_metrics['Standard Deviation']

        # Set up the plot canvas
        fig, ax = plt.subplots(figsize=(16, 10))
        fig.patch.set_facecolor('white')     # Set white background
        ax.set_facecolor('lightgrey')        # Set plot area to light grey
        ax.grid(True)                        # Show grid lines for easier reading

        # Scatter plot: x = volatility, y = return, color = Sharpe ratio
        scatter = ax.scatter(
            self.results[1],                    # Standard deviation (risk) on x-axis
            self.results[0] + 1,                # Cumulative return + 1 on y-axis
            c=self.results[2],                  # Sharpe ratio determines color
            cmap='RdYlBu',                      # Color map from red to blue
            marker='o',
            edgecolor='k',
            alpha=0.8,
            s=30
        )

        plt.colorbar(scatter, label='Sharpe Ratio')     # Add legend for color scale
        ax.set_title('Efficient Frontier')              # Plot title
        ax.set_xlabel('Volatility (Standard Deviation)')  # x-axis label
        ax.set_ylabel('Portfolio Return [x times]')     # y-axis label

        # Add marker for the Sharpe-optimal portfolio
        ax.scatter(
            optimal_risk,
            optimal_return,
            color='green',
            marker='*',
            s=400,
            label='Optimal Risk Adjusted Portfolio'
        )

        # Add marker for the lowest-risk portfolio in the simulation
        min_risk_idx = np.argmin(self.results[1])       # Index of portfolio with lowest std dev
        min_risk_return = self.results[0, min_risk_idx] + 1
        min_risk_volatility = self.results[1, min_risk_idx]
        ax.scatter(
            min_risk_volatility,
            min_risk_return,
            color='blue',
            marker='*',
            s=300,
            label='Least Risky Portfolio'
        )

        ax.legend()  # Show legend for the markers

        # Add explanation text to bottom of chart
        ax.text(
            0, -0.2,
            'The Sharpe Ratio measures the risk-adjusted return of a portfolio.\n'
            'The optimal portfolio maximizes the Sharpe Ratio,\nproviding the best trade-off between return and risk.',
            transform=ax.transAxes,
            fontsize=10,
            verticalalignment='bottom',
            horizontalalignment='left',
            bbox=dict(facecolor='white', alpha=0.5)
        )

        # Add interactive cursor for hover tooltips
        cursor = mplcursors.cursor(scatter, hover=True)

        @cursor.connect("add")
        def on_hover(sel):
            index = sel.index
            sel.annotation.set_text(
                f"Weights: {self.weights[index]}\n"
                f"Return: {(self.results[0, index] + 1):.2f}X\n"
                f"Std Dev: {self.results[1, index]:.2f}"
            )
            sel.annotation.get_bbox_patch().set(facecolor='black', alpha=0.75)
            sel.annotation.set_color('white')

        # Adjust layout and save to file
        plt.tight_layout()
        plt.savefig(save_path)
        plt.show()


class ReportGenerator:
    def __init__(
        self,
        etf_data_handler: object,
        etf_metrics: object,
        optimizer: object,
        results: pd.DataFrame,
        plot_path: str = 'efficient_frontier.png'
    ) -> None:
        """
        Initialize the report generator with all required data components.

        Parameters:
            etf_data_handler (object): Object containing raw ETF data.
            etf_metrics (object): Object containing ETF-level statistics and correlation.
            optimizer (object): PortfolioOptimizer instance with optimal weights and metrics.
            results (pd.DataFrame): Full simulation results matrix.
            plot_path (str): Filepath to efficient frontier image for embedding.
        """
        self.etf_data = etf_data_handler.data
        self.etf_metrics = etf_metrics
        self.optimal_weights = optimizer.optimal_weights
        self.optimal_metrics = optimizer.optimal_metrics
        self.results = results
        self.plot_path = plot_path
        self.constraints = {
            'Max ETF Weight': optimizer.max_etf_weight,
            'Weight Step Size': optimizer.weight_step_size
        }

    def prepare_report_data(self) -> dict:
        """
        Prepare the dictionary of dataframes and metadata for Excel output.

        Returns:
            dict: Mapping of sheet names to dataframes for report generation.
        """
        report_data = {
            'Historical ETF Data': self.etf_data,
            'ETF Metrics': self.etf_metrics.metrics.copy(),  # Clone to avoid inplace edits
            'Correlation Matrix': self.etf_metrics.correlation_stats['Correlation Matrix'],
            'Rolling Correlation (30-day)': self.etf_metrics.correlation_stats['Rolling Correlation (30-day)'],
            'Mean Correlation': self.etf_metrics.correlation_stats['Mean Correlation'],
            'Median Correlation': self.etf_metrics.correlation_stats['Median Correlation'],
            'Optimized Portfolio Metrics': pd.DataFrame([self.optimal_metrics]),
            'Constraints': self.constraints
        }

        # Append optimal weights directly into the ETF metrics table
        report_data['ETF Metrics']['Optimal Weights'] = self.optimal_weights

        # Remove redundant Optimal Weights from the summary tab (kept separately above)
        del report_data['Optimized Portfolio Metrics']['Optimal Weights']

        return report_data

    def get_report_filename(self) -> str:
        """
        Generate a descriptive filename based on ETFs and current date.

        Returns:
            str: Auto-generated report filename.
        """
        portfolio_str = '_'.join(self.etf_data.columns)
        current_date = datetime.now().strftime('%Y-%m-%d')
        return f'etf_report_{portfolio_str}_{current_date}.xlsx'

    def save_report(self, report_data: dict, filename: str | None = None) -> None:
        """
        Save all data and plots to a formatted Excel workbook.

        Parameters:
            report_data (dict): Output of prepare_report_data()
            filename (str, optional): Custom filename. If None, auto-generated.
        """
        if filename is None:
            filename = self.get_report_filename()

        # Create the base Excel workbook
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            report_data['Historical ETF Data'].to_excel(writer, sheet_name='Historical ETF Data', index=True)
            report_data['ETF Metrics'].to_excel(writer, sheet_name='ETF Metrics', index=True)
            report_data['Correlation Matrix'].to_excel(writer, sheet_name='Correlation Matrix', index=True)
            report_data['Rolling Correlation (30-day)'].to_excel(writer, sheet_name='Rolling Correlation (30-day)', index=False)
            report_data['Optimized Portfolio Metrics'].to_excel(writer, sheet_name='Optimized Portfolio Metrics', index=True)

        # Load workbook again to modify it (column widths + add image)
        workbook = load_workbook(filename)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Auto-adjust column widths based on max text length in each column
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter

                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        continue

                worksheet.column_dimensions[col_letter].width = max_length

        # Embed the plot image into the summary sheet
        if 'Optimized Portfolio Metrics' in workbook.sheetnames:
            worksheet = workbook['Optimized Portfolio Metrics']
            img = Image(self.plot_path)
            img.anchor = "B4"  # Top-left cell for the image
            worksheet.add_image(img)

        # Get the directory of the current script
        current_dir = os.path.dirname(os.path.abspath(__file__))

        # Save final workbook with image and adjustments
        file_path = os.path.join(current_dir, 'Reports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        workbook.save(file_path)
        print(f"\nSaving to: \n{file_path}")


    def condition_metrics_for_percentage(self) -> None:
        """
        Convert absolute return values to human-readable percentages for Excel display.
        """
        # Format ETF-level metrics
        self.etf_metrics.metrics['Real Return (%)'] = self.etf_metrics.metrics['Real Return'] * 100
        self.etf_metrics.metrics['CAGR (%)'] = self.etf_metrics.metrics['Annualized Return'] * 100
        self.etf_metrics.metrics = self.etf_metrics.metrics.drop(columns=['Real Return', 'Annualized Return'])

        # Format portfolio-level metrics
        self.optimal_metrics['Real Return (%)'] = self.optimal_metrics['Real Return'] * 100
        self.optimal_metrics['CAGR (%)'] = self.optimal_metrics['Annualized Return'] * 100
        del self.optimal_metrics['Real Return']
        del self.optimal_metrics['Annualized Return']



# Main function to execute the process
def main():
    asset_list = [
        ['SMH', 'IHI', 'XBI', 'ITB', 'SOXL'],  # Line 0
        ['SPY', 'QQQ', 'VTI', 'VWO', 'VEA', 'QUAL'],  # Line 1, etc.
        ["TAN", "SMH", "ARGT", "IYW", 'IHI'],  # 2
        ["SOXL", "SMH", "ARGT", "IYW", 'IHI'],  # 3
        ["QUAL", "SMH", "ARGT", "IYW", 'IHI'],  # 4
        ["QUAL", "SMH", "ARGT", "IYW", 'IHI', 'SOXL'],  # 5
        ["EEM", "SMH", "ARGT", "IYW", 'IHI', 'QUAL'],  # 6
        ["TAN", "SMH", "ARGT", "IYW", "SPY", "EFA",
            "EEM", "AGG", "GLD", "VNQ"],  # 7
        ['SMH', 'IHI', 'XBI', 'QUAL', 'SOXL'],  # 8
        ['SMH', 'IHI', 'XBI', 'ITB', 'SOXL'],  # 9
        ['SMH', 'IHI', 'HODL', 'SOXL', 'QUAL'],  # 10
        ['SMH', 'IHI', 'ITB', 'SOXL', 'QUAL'],  # 11
        ['IYW', 'COPX', 'SPUU', 'XLK', 'SSO', 'ITB', 'AIRR', 'FTEC', 'XHB', 'VGT',
            'FCG', 'IXN', 'XSD', 'PAVE', 'GRN', 'PTF', 'IGM', 'GRID', 'QQQ', 'XME'],  # 12
        ['IYW', 'COPX', 'XLK', 'VGT', 'FTEC'],  # 13
        ['IYW', 'COPX', 'SPUU', 'SSO', 'GRN'],  # 14
        ['IVV', 'TDT.AS', 'EWD', 'QUAL', 'SOXL'],  # 15
        ['SSO', 'PAVE', 'XHB', 'GRID', 'XME', 'IYW', 'COPX'],  # 16
        ['AAPL', 'AMZN', 'MSFT'],  # 17
        ['AAPL', 'AMZN'],
        ['AAPL'],
        ['AAPL', 'AMZN', 'MSFT', 'GOOGL', 'TSLA',
            'NVDA', 'DIS', 'NFLX', 'CMCSA', 'T'],
        ["SPUU", "FTEC", "ARGT", "IYW", 'SSO', 'GRN'],  # 21
        ["FTEC", "ITB", "ARGT"],  # 22
        ["FTEC", "ITB", "ARGT", "PPA"],  # 23
        ["FTEC", "ITB", "ARGT", "ITA"],  # 24
    ]

    # Select one of the sets to experiment with
    symbols = asset_list[24]
    start_date = '2015-06-29'   # American date format, e.g., 2015-06-29
    end_date = '2025-06-29'
    # end_date = pd.Timestamp.today().strftime('%Y-%m-%d')
    # Step 1: Download Data
    etf_data_handler = ETFDataHandler(symbols, start_date, end_date)

    # Step 2: Calculate Metrics
    etf_metrics = ETFMetrics(etf_data_handler)

    # Display Correlation Statistics
    etf_metrics.display_correlation_statistics()

    # Step 3: Optimize Portfolio
    risk_free_rate = 0.02  # Set the risk-free rate (e.g., 2%)
    optimizer = PortfolioOptimizer(etf_metrics, risk_free_rate)

    # Step 4: Plot Efficient Frontier
    plotter = EfficientFrontierPlotter(optimizer)
    plotter.plot_efficient_frontier(save_path='efficient_frontier.png')

    # Step 5: Generate Report
    report_generator = ReportGenerator(
        etf_data_handler, etf_metrics, optimizer, optimizer.results_matrix, 'efficient_frontier.png')
    report_generator.condition_metrics_for_percentage()
    report_data = report_generator.prepare_report_data()
    report_generator.save_report(report_data, filename=None)
    print("Exiting program")


if __name__ == "__main__":
    main()

# Step 1: Include the symbols you'd like to analyse in the asset_list = [] list. Every line (= ['ETF1', 'ETF2'],) is a portfolio.
# Step 2:  symbols = asset_list[0], 0 indicates it will analyse row 0 of the list that you populated in step 1. Likewise, 1 means the programme will analyse the portfolio you defined on line 1, etc.
# Step 3: OPTIONAL: self.max_etf_weight = 0.3  # Maximum weight for a single ETF. 0.3 corresponds to 30%, meaning a single ETF of the portfolio you defined in step 1, cannot exceed a maximum weight of 30%.
# and self.weight_step_size = 1  # Weight step size in increments of 1%. The programme calculated returns and risk for every possible combination of portfolio compositions in steps of 1%. ALterantively, you can set it to larger steps.
# Step 4: Set the risk-free rate

# To do:
# request user where to save the report.
# Suggest more descriptive file name: Porfolio analysis ETF1 ETF2 etc.xslx
# add check to see if ETF ticker exists. If not, prompt user to enter ticker correctly OR skip ticker.
# Add check to determine if there is data for ALL ETF's for the stated period. If NOT:
    # determine the longest period available. Find MAX of all ETF's and compare.

# Add top 10 of max sharpe ratio portfolios to analysis sheet + lowest risk + highest return portfolio, inculding their metrics.
# Create GUI, such that the user can reload the portfolio and hover over the graph to see all sorts of portfolio compositions, their returns and risks.
#
# Create executable file.
